import mysql.connector
import requests
import openpyxl

# Open the workbook and define the worksheet
book = openpyxl.load_workbook("HINDALCO_1D.xlsx")
sheet = book.active

# Establish a MySQL connection
database =  mysql.connector.connect (host="localhost", user = "root", passwd = "Ritik@MySql123", db = "trading_algo")

# Get the cursor, which is used to traverse the database, line by line
cursor = database.cursor()

# Create the INSERT INTO sql query
query = """INSERT INTO hindalco (Date_Time, Close, High, Low, Open, Volume, Instrument) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

# Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
for r in range(2, sheet.max_row+1):
		Date_Time	= sheet.cell(r,1).value
		Close		= sheet.cell(r,2).value
		High 	    = sheet.cell(r,3).value
		Low  		= sheet.cell(r,4).value
		Open    	= sheet.cell(r,5).value
		Volume		= sheet.cell(r,6).value
		Instrument	= sheet.cell(r,7).value

		# Assign values from each row
		values = (Date_Time, Close, High, Low, Open, Volume, Instrument)
		# Execute sql Query
		cursor.execute(query, values)


# Close the cursor
cursor.close()

# Commit the transaction
database.commit()

# Close the database connection
database.close()

# Print results
print ("Data Successfully Added to database !!")

columns = str(sheet.max_column)
rows = str(sheet.max_row)
print("Rows   : ", rows)
print("Column : ", columns)